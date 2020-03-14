from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv
def parse_csv_file(file:str):
    list = []
    with open(file) as f:
        lines = csv.reader(f)
        next(lines)  # 去掉表头
        for line in lines:
            # print(tuple(line))
            list.append(tuple(line))
            # pass
    return list

def parse_excel_file(filename,sheetname):

    workbook = load_workbook(filename)
    ws:Worksheet = workbook[sheetname]

    testdata = []
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column,values_only=True):
        testdata.append(row)

    return testdata